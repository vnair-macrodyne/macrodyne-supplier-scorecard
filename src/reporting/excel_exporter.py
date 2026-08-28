import json
from pathlib import Path

import pandas as pd

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ==================================================
# COLOURS / STYLE CONSTANTS
# ==================================================

DARK_BLUE = "17365D"
MEDIUM_BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"

DARK_GREEN = "548235"
LIGHT_GREEN = "E2F0D9"

GOLD = "BF9000"
LIGHT_GOLD = "FFF2CC"

ORANGE = "C65911"
LIGHT_ORANGE = "FCE4D6"

RED = "C00000"
LIGHT_RED = "F4CCCC"

GRAY = "7F7F7F"
LIGHT_GRAY = "E7E6E6"

WHITE = "FFFFFF"
BLACK = "000000"


THIN_GRAY = Side(
    style="thin",
    color="D9D9D9"
)

STANDARD_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY
)


# ==================================================
# HELPER - LOAD SCORECARD RULES
# ==================================================

def _load_scorecard_rules(
    rules_path="config/scorecard_rules.json"
):
    """
    Load the current prototype scorecard configuration.

    The exporter uses this only for documenting the
    assumptions in the Prototype Notes worksheet.
    """

    rules_file = Path(rules_path)

    if not rules_file.exists():
        return None

    with open(
        rules_file,
        "r",
        encoding="utf-8"
    ) as config_file:

        return json.load(
            config_file
        )


# ==================================================
# HELPER - SAFE COLUMN SELECTION
# ==================================================

def _select_existing_columns(
    dataframe,
    column_mapping
):
    """
    Select only columns that currently exist.

    column_mapping:
        {
            internal_column: display_column
        }

    This makes the exporter more resilient if a future
    prototype version temporarily removes a field.
    """

    existing_mapping = {
        source_column: display_column
        for source_column, display_column
        in column_mapping.items()
        if source_column in dataframe.columns
    }

    export_df = dataframe[
        list(
            existing_mapping.keys()
        )
    ].copy()

    export_df = export_df.rename(
        columns=existing_mapping
    )

    return export_df


# ==================================================
# HELPER - GENERAL SHEET FORMATTING
# ==================================================

def _format_standard_sheet(
    worksheet,
    header_row,
    freeze_cell=None
):
    """
    Apply standard formatting to a normal data worksheet.
    """

    max_column = worksheet.max_column
    max_row = worksheet.max_row

    # --------------------------------------------------
    # HEADER
    # --------------------------------------------------

    for cell in worksheet[
        header_row
    ]:

        cell.fill = PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )

        cell.font = Font(
            color=WHITE,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = STANDARD_BORDER


    # --------------------------------------------------
    # BODY
    # --------------------------------------------------

    for row in worksheet.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_column
    ):

        for cell in row:

            cell.border = STANDARD_BORDER

            cell.alignment = Alignment(
                vertical="top"
            )


    # --------------------------------------------------
    # FREEZE / FILTER
    # --------------------------------------------------

    if freeze_cell:
        worksheet.freeze_panes = freeze_cell

    worksheet.auto_filter.ref = (
        f"A{header_row}:"
        f"{get_column_letter(max_column)}"
        f"{max_row}"
    )


# ==================================================
# HELPER - GRADE CELL COLOUR
# ==================================================

def _apply_grade_style(
    cell
):
    """
    Colour an A/B/C/D/N/A grade cell.
    """

    grade = str(
        cell.value
    ).strip().upper()


    if grade == "A":

        cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_GREEN
        )

        cell.font = Font(
            bold=True,
            color=DARK_GREEN
        )


    elif grade == "B":

        cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_BLUE
        )

        cell.font = Font(
            bold=True,
            color=DARK_BLUE
        )


    elif grade == "C":

        cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_GOLD
        )

        cell.font = Font(
            bold=True,
            color=GOLD
        )


    elif grade == "D":

        cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_RED
        )

        cell.font = Font(
            bold=True,
            color=RED
        )


    else:

        cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_GRAY
        )

        cell.font = Font(
            bold=True,
            color=GRAY
        )


    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    cell.border = STANDARD_BORDER


# ==================================================
# HELPER - COLUMN WIDTHS
# ==================================================

def _set_reasonable_widths(
    worksheet,
    maximum_width=32
):
    """
    Estimate readable column widths while preventing
    extremely wide columns.
    """

    for column_cells in worksheet.columns:

        column_letter = (
            get_column_letter(
                column_cells[0].column
            )
        )

        if worksheet.column_dimensions[
            column_letter
        ].hidden:
            continue


        maximum_length = 0

        for cell in column_cells:

            if cell.value is None:
                continue

            value_length = len(
                str(
                    cell.value
                )
            )

            maximum_length = max(
                maximum_length,
                value_length
            )


        adjusted_width = min(
            maximum_length + 2,
            maximum_width
        )

        adjusted_width = max(
            adjusted_width,
            10
        )

        worksheet.column_dimensions[
            column_letter
        ].width = adjusted_width


# ==================================================
# HELPER - FIND EXCEL COLUMN
# ==================================================

def _header_column_map(
    worksheet,
    header_row
):
    """
    Build:
        Display Header -> Excel Column Letter
    """

    mapping = {}

    for cell in worksheet[
        header_row
    ]:

        if cell.value is None:
            continue

        mapping[
            str(
                cell.value
            )
        ] = get_column_letter(
            cell.column
        )

    return mapping


# ==================================================
# VENDOR SCORECARD SHEET
# ==================================================

def _build_vendor_scorecard_sheet(
    writer,
    vendor_summary
):
    """
    Create the main 416-vendor scorecard worksheet.
    """

    scorecard = vendor_summary.copy()


    # ==================================================
    # VENDOR KEY
    # ==================================================

    scorecard[
        "vendor_key"
    ] = (
        scorecard[
            "vendor_match_name"
        ]
        .fillna(
            "UNKNOWN VENDOR"
        )
        .astype(str)
        +
        " | "
        +
        scorecard[
            "vendor_match_city"
        ]
        .fillna(
            "NO LOCATION"
        )
        .astype(str)
    )


    # ==================================================
    # OUTPUT COLUMN ORDER
    # ==================================================

    column_mapping = {

        # ----------------------------------------------
        # IDENTIFICATION
        # ----------------------------------------------

        "vendor_key":
            "Vendor Key",

        "vendor_match_name":
            "Vendor",

        "vendor_match_city":
            "Location",


        # ----------------------------------------------
        # OVERALL SCORE
        # ----------------------------------------------

        "prototype_overall_score":
            "Prototype Overall Score",

        "prototype_overall_grade":
            "Prototype Overall Grade",

        "prototype_weight_coverage_pct":
            "Weight Coverage %",

        "prototype_scored_component_count":
            "Scored Components",

        "prototype_overall_status":
            "Overall Status",


        # ----------------------------------------------
        # ACTIVITY
        # ----------------------------------------------

        "po_transaction_count":
            "PO Transactions",

        "distinct_po_count":
            "Distinct POs",

        "total_ordered_qty":
            "Total Ordered Qty",

        "total_received_qty":
            "Total Received Qty",


        # ----------------------------------------------
        # DELIVERY
        # ----------------------------------------------

        "delivery_eligible_count":
            "Delivery Eligible",

        "on_time_count":
            "On-Time",

        "late_count":
            "Late",

        "on_time_delivery_pct":
            "OTD %",

        "average_days_late":
            "Average Days Late",

        "delivery_prototype_score":
            "Delivery Score",

        "delivery_prototype_grade":
            "Delivery Grade",

        "delivery_score_status":
            "Delivery Status",


        # ----------------------------------------------
        # QUALITY
        # ----------------------------------------------

        "supplier_linked_ncr_count":
            "Supplier-Linked NCRs",

        "supplier_linked_ncr_rate_pct":
            "Supplier-Linked NCR Rate %",

        "quality_eligible_ncr_count":
            "Quality-Eligible NCRs",

        "total_ncr_quantity":
            "Valid NCR Quantity",

        "quality_rejected_qty":
            "Valid Rejected Qty",

        "ncr_rejected_pct":
            "NCR Rejected %",

        "ncr_quantity_anomaly_count":
            "NCR Qty Anomalies",

        "quality_prototype_score":
            "Quality Score",

        "quality_prototype_grade":
            "Quality Grade",

        "quality_score_status":
            "Quality Status",


        # ----------------------------------------------
        # LEAD TIME
        # ----------------------------------------------

        "lead_time_eligible_count":
            "Lead-Time Eligible",

        "lead_time_adherent_count":
            "Lead-Time Adherent",

        "average_actual_lead_time_days":
            "Average Actual Lead Time",

        "average_lead_time_variance_days":
            "Average Lead-Time Variance",

        "lead_time_adherence_pct":
            "Lead-Time Adherence %",

        "lead_time_prototype_score":
            "Lead-Time Score",

        "lead_time_prototype_grade":
            "Lead-Time Grade",

        "lead_time_score_status":
            "Lead-Time Status",


        # ----------------------------------------------
        # RESPONSIVENESS
        # ----------------------------------------------

        "responsiveness_eligible_ncr_count":
            "Responsiveness-Eligible NCRs",

        "resolved_ncr_count":
            "Resolved NCRs",

        "unresolved_ncr_count":
            "Unresolved NCRs",

        "responsiveness_proxy_pct":
            "Responsiveness Proxy - NCR Resolution %",

        "responsiveness_prototype_score":
            "Responsiveness Score",

        "responsiveness_prototype_grade":
            "Responsiveness Grade",

        "responsiveness_score_status":
            "Responsiveness Status",


        # ----------------------------------------------
        # COMMERCIAL
        # ----------------------------------------------

        "price_comparison_count":
            "Price Comparisons",

        "price_stable_count":
            "Price Stable / Decreased",

        "price_increase_count":
            "Price Increased",

        "average_price_change_pct":
            "Average Price Change %",

        "price_stability_pct":
            "Price Stability %",

        "commercial_prototype_score":
            "Commercial Score",

        "commercial_prototype_grade":
            "Commercial Grade",

        "commercial_score_status":
            "Commercial Status"
    }


    scorecard_export = (
        _select_existing_columns(
            scorecard,
            column_mapping
        )
    )


    scorecard_export = (
        scorecard_export
        .sort_values(
            by=[
                "Vendor",
                "Location"
            ],
            na_position="last"
        )
        .reset_index(
            drop=True
        )
    )


    # ==================================================
    # EXPORT
    # ==================================================

    scorecard_export.to_excel(
        writer,
        sheet_name="Vendor Scorecard",
        index=False,
        startrow=3
    )


    worksheet = (
        writer.book[
            "Vendor Scorecard"
        ]
    )


    # ==================================================
    # TITLE
    # ==================================================

    last_column = (
        get_column_letter(
            scorecard_export.shape[1]
        )
    )


    worksheet.merge_cells(
        f"A1:{last_column}1"
    )

    worksheet["A1"] = (
        "Macrodyne Vendor Scorecard Prototype"
    )

    worksheet["A1"].fill = (
        PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )
    )

    worksheet["A1"].font = Font(
        color=WHITE,
        bold=True,
        size=16
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.row_dimensions[1].height = 28


    worksheet.merge_cells(
        f"A2:{last_column}2"
    )

    worksheet["A2"] = (
        "PROTOTYPE ONLY — weights, thresholds, "
        "sample rules and grades require business approval."
    )

    worksheet["A2"].fill = (
        PatternFill(
            "solid",
            fgColor=LIGHT_GOLD
        )
    )

    worksheet["A2"].font = Font(
        bold=True,
        color=GOLD
    )

    worksheet["A2"].alignment = Alignment(
        horizontal="center"
    )


    # ==================================================
    # HEADER
    # ==================================================

    header_row = 4

    _format_standard_sheet(
        worksheet,
        header_row=header_row,
        freeze_cell="D5"
    )


    header_map = (
        _header_column_map(
            worksheet,
            header_row
        )
    )


    # ==================================================
    # NUMBER FORMATS
    # ==================================================

    percentage_headers = [
        "Weight Coverage %",
        "OTD %",
        "Supplier-Linked NCR Rate %",
        "NCR Rejected %",
        "Lead-Time Adherence %",
        "Responsiveness Proxy - NCR Resolution %",
        "Average Price Change %",
        "Price Stability %"
    ]

    score_headers = [
        "Prototype Overall Score",
        "Delivery Score",
        "Quality Score",
        "Lead-Time Score",
        "Responsiveness Score",
        "Commercial Score"
    ]


    first_data_row = 5
    last_data_row = worksheet.max_row


    for header in percentage_headers:

        if header not in header_map:
            continue

        column_letter = (
            header_map[
                header
            ]
        )

        for cell in worksheet[
            f"{column_letter}"
            f"{first_data_row}:"
            f"{column_letter}"
            f"{last_data_row}"
        ]:

            cell[0].number_format = (
                '0.0\\%'
            )


    for header in score_headers:

        if header not in header_map:
            continue

        column_letter = (
            header_map[
                header
            ]
        )

        for cell in worksheet[
            f"{column_letter}"
            f"{first_data_row}:"
            f"{column_letter}"
            f"{last_data_row}"
        ]:

            cell[0].number_format = (
                "0.0"
            )


    # ==================================================
    # GRADE FORMATTING
    # ==================================================

    grade_headers = [
        "Prototype Overall Grade",
        "Delivery Grade",
        "Quality Grade",
        "Lead-Time Grade",
        "Responsiveness Grade",
        "Commercial Grade"
    ]


    for header in grade_headers:

        if header not in header_map:
            continue

        column_letter = (
            header_map[
                header
            ]
        )

        for row_number in range(
            first_data_row,
            last_data_row + 1
        ):

            _apply_grade_style(
                worksheet[
                    f"{column_letter}"
                    f"{row_number}"
                ]
            )


    # ==================================================
    # EMPHASIZE OVERALL COLUMNS
    # ==================================================

    overall_headers = [
        "Prototype Overall Score",
        "Prototype Overall Grade",
        "Weight Coverage %",
        "Scored Components",
        "Overall Status"
    ]


    for header in overall_headers:

        if header not in header_map:
            continue

        column_letter = (
            header_map[
                header
            ]
        )

        worksheet[
            f"{column_letter}4"
        ].fill = PatternFill(
            "solid",
            fgColor=DARK_GREEN
        )


    # ==================================================
    # HIDE INTERNAL VENDOR KEY
    # ==================================================

    if "Vendor Key" in header_map:

        worksheet.column_dimensions[
            header_map[
                "Vendor Key"
            ]
        ].hidden = True


    _set_reasonable_widths(
        worksheet,
        maximum_width=30
    )


    # Give Vendor more room.
    if "Vendor" in header_map:

        worksheet.column_dimensions[
            header_map[
                "Vendor"
            ]
        ].width = 38


    return (
        worksheet,
        scorecard_export
    )


# ==================================================
# VENDOR DETAIL SHEET
# ==================================================

def _build_vendor_detail_sheet(
    writer,
    scorecard_export,
    scorecard_sheet
):
    """
    Build an interactive individual-vendor scorecard.

    Cell B3 contains a dropdown.

    The displayed values use INDEX/MATCH against the
    Vendor Scorecard worksheet so selecting another vendor
    automatically refreshes the scorecard in Excel.
    """

    workbook = writer.book

    worksheet = (
        workbook.create_sheet(
            "Vendor Detail"
        )
    )


    # ==================================================
    # TITLE
    # ==================================================

    worksheet.merge_cells(
        "A1:G1"
    )

    worksheet["A1"] = (
        "Individual Vendor Scorecard"
    )

    worksheet["A1"].fill = (
        PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )
    )

    worksheet["A1"].font = Font(
        color=WHITE,
        bold=True,
        size=16
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    worksheet.row_dimensions[1].height = 28


    # ==================================================
    # VENDOR SELECTOR
    # ==================================================

    worksheet["A3"] = (
        "Select Vendor / Location"
    )

    worksheet["A3"].font = Font(
        bold=True,
        color=DARK_BLUE
    )

    worksheet["B3"] = ""


    # Hidden vendor list in column Z.
    vendor_keys = (
        scorecard_export[
            "Vendor Key"
        ]
        .dropna()
        .drop_duplicates()
        .tolist()
    )


    for index, vendor_key in enumerate(
        vendor_keys,
        start=2
    ):

        worksheet[
            f"Z{index}"
        ] = vendor_key


    worksheet.column_dimensions[
        "Z"
    ].hidden = True


    # Default to highest scored vendor when possible.
    scored_vendors = (
        scorecard_export.loc[
            scorecard_export[
                "Prototype Overall Score"
            ].notna()
        ]
        .sort_values(
            "Prototype Overall Score",
            ascending=False
        )
    )


    if not scored_vendors.empty:

        worksheet["B3"] = (
            scored_vendors.iloc[0][
                "Vendor Key"
            ]
        )

    elif vendor_keys:

        worksheet["B3"] = (
            vendor_keys[0]
        )


    if vendor_keys:

        validation = DataValidation(
            type="list",
            formula1=(
                f"$Z$2:"
                f"$Z${len(vendor_keys) + 1}"
            ),
            allow_blank=False
        )

        worksheet.add_data_validation(
            validation
        )

        validation.add(
            worksheet[
                "B3"
            ]
        )


    worksheet["B3"].fill = (
        PatternFill(
            "solid",
            fgColor=LIGHT_BLUE
        )
    )

    worksheet["B3"].font = Font(
        bold=True
    )


    # ==================================================
    # SUMMARY HEADER
    # ==================================================

    worksheet.merge_cells(
        "A5:G5"
    )

    worksheet["A5"] = (
        "Prototype Overall Assessment"
    )

    worksheet["A5"].fill = (
        PatternFill(
            "solid",
            fgColor=DARK_GREEN
        )
    )

    worksheet["A5"].font = Font(
        color=WHITE,
        bold=True
    )

    worksheet["A5"].alignment = Alignment(
        horizontal="center"
    )


    summary_labels = [
        ("A6", "Vendor"),
        ("A7", "Location"),
        ("C6", "PO Transactions"),
        ("C7", "Scored Components"),
        ("E6", "Overall Score"),
        ("E7", "Overall Grade"),
        ("A8", "Weight Coverage %"),
        ("C8", "Overall Status")
    ]


    for (
        cell_address,
        label
    ) in summary_labels:

        worksheet[
            cell_address
        ] = label

        worksheet[
            cell_address
        ].font = Font(
            bold=True,
            color=DARK_BLUE
        )


    # ==================================================
    # COMPONENT TABLE
    # ==================================================

    component_start_row = 11

    component_headers = [
        "Component",
        "Configured Weight",
        "Primary Metric",
        "Sample",
        "Prototype Score",
        "Grade",
        "Status"
    ]


    for column_index, header in enumerate(
        component_headers,
        start=1
    ):

        cell = worksheet.cell(
            row=component_start_row,
            column=column_index
        )

        cell.value = header

        cell.fill = PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )

        cell.font = Font(
            color=WHITE,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = STANDARD_BORDER


    component_rows = {

        12: (
            "On-Time Delivery",
            "25%",
            "OTD %",
            "Delivery Eligible",
            "Delivery Score",
            "Delivery Grade",
            "Delivery Status"
        ),

        13: (
            "Quality / NCR",
            "25%",
            "Supplier-Linked NCR Rate %",
            "PO Transactions",
            "Quality Score",
            "Quality Grade",
            "Quality Status"
        ),

        14: (
            "Lead-Time Performance",
            "15%",
            "Lead-Time Adherence %",
            "Lead-Time Eligible",
            "Lead-Time Score",
            "Lead-Time Grade",
            "Lead-Time Status"
        ),

        15: (
            "Responsiveness Proxy",
            "15%",
            "Responsiveness Proxy - NCR Resolution %",
            "Responsiveness-Eligible NCRs",
            "Responsiveness Score",
            "Responsiveness Grade",
            "Responsiveness Status"
        ),

        16: (
            "Commercial Performance",
            "20%",
            "Price Stability %",
            "Price Comparisons",
            "Commercial Score",
            "Commercial Grade",
            "Commercial Status"
        )
    }


    # ==================================================
    # DETAIL SECTIONS
    # ==================================================

    worksheet.merge_cells(
        "A19:G19"
    )

    worksheet["A19"] = (
        "Supporting Metrics"
    )

    worksheet["A19"].fill = (
        PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )
    )

    worksheet["A19"].font = Font(
        color=WHITE,
        bold=True
    )

    worksheet["A19"].alignment = Alignment(
        horizontal="center"
    )


    detail_metrics = [

        # ----------------------------------------------
        # DELIVERY
        # ----------------------------------------------

        (
            21,
            "Delivery",
            [
                (
                    "Delivery Eligible",
                    "Delivery Eligible"
                ),
                (
                    "On-Time",
                    "On-Time"
                ),
                (
                    "Late",
                    "Late"
                ),
                (
                    "OTD %",
                    "OTD %"
                ),
                (
                    "Average Days Late",
                    "Average Days Late"
                )
            ]
        ),


        # ----------------------------------------------
        # QUALITY
        # ----------------------------------------------

        (
            28,
            "Quality / NCR",
            [
                (
                    "Supplier-Linked NCRs",
                    "Supplier-Linked NCRs"
                ),
                (
                    "Supplier-Linked NCR Rate %",
                    "Supplier-Linked NCR Rate %"
                ),
                (
                    "Quality-Eligible NCRs",
                    "Quality-Eligible NCRs"
                ),
                (
                    "NCR Rejected %",
                    "NCR Rejected %"
                ),
                (
                    "NCR Qty Anomalies",
                    "NCR Qty Anomalies"
                )
            ]
        ),


        # ----------------------------------------------
        # LEAD TIME
        # ----------------------------------------------

        (
            35,
            "Lead-Time Performance",
            [
                (
                    "Lead-Time Eligible",
                    "Lead-Time Eligible"
                ),
                (
                    "Lead-Time Adherent",
                    "Lead-Time Adherent"
                ),
                (
                    "Lead-Time Adherence %",
                    "Lead-Time Adherence %"
                ),
                (
                    "Average Actual Lead Time",
                    "Average Actual Lead Time"
                ),
                (
                    "Average Lead-Time Variance",
                    "Average Lead-Time Variance"
                )
            ]
        ),


        # ----------------------------------------------
        # RESPONSIVENESS
        # ----------------------------------------------

        (
            42,
            "Responsiveness Proxy",
            [
                (
                    "Responsiveness-Eligible NCRs",
                    "Responsiveness-Eligible NCRs"
                ),
                (
                    "Resolved NCRs",
                    "Resolved NCRs"
                ),
                (
                    "Unresolved NCRs",
                    "Unresolved NCRs"
                ),
                (
                    "NCR Resolution %",
                    "Responsiveness Proxy - NCR Resolution %"
                )
            ]
        ),


        # ----------------------------------------------
        # COMMERCIAL
        # ----------------------------------------------

        (
            48,
            "Commercial Performance",
            [
                (
                    "Price Comparisons",
                    "Price Comparisons"
                ),
                (
                    "Price Stable / Decreased",
                    "Price Stable / Decreased"
                ),
                (
                    "Price Increased",
                    "Price Increased"
                ),
                (
                    "Average Price Change %",
                    "Average Price Change %"
                ),
                (
                    "Price Stability %",
                    "Price Stability %"
                )
            ]
        )
    ]


    # ==================================================
    # LOOKUP SETUP
    # ==================================================

    scorecard_header_row = 4

    scorecard_header_map = (
        _header_column_map(
            scorecard_sheet,
            scorecard_header_row
        )
    )

    scorecard_last_row = (
        scorecard_sheet.max_row
    )


    key_column = (
        scorecard_header_map[
            "Vendor Key"
        ]
    )


    def lookup_formula(
        display_header
    ):
        """
        Return an INDEX/MATCH formula using Vendor Key.

        Blank source cells are displayed as N/A instead
        of Excel converting them to numeric zero.
        """

        if (
            display_header
            not in scorecard_header_map
        ):
            return '="N/A"'


        value_column = (
            scorecard_header_map[
                display_header
            ]
        )


        value_range = (
            f"'Vendor Scorecard'!"
            f"${value_column}$5:"
            f"${value_column}${scorecard_last_row}"
        )


        key_range = (
            f"'Vendor Scorecard'!"
            f"${key_column}$5:"
            f"${key_column}${scorecard_last_row}"
        )


        match_formula = (
            f"MATCH("
            f"$B$3,"
            f"{key_range},"
            f"0"
            f")"
        )


        index_formula = (
            f"INDEX("
            f"{value_range},"
            f"{match_formula}"
            f")"
        )


        return (
            "=IFERROR("
            "IF("
            f"ISBLANK({index_formula}),"
            '"N/A",'
            f"{index_formula}"
            "),"
            '"N/A"'
            ")"
        )
        """
        Return an INDEX/MATCH formula using Vendor Key.
        """

        if (
            display_header
            not in scorecard_header_map
        ):
            return '="N/A"'


        value_column = (
            scorecard_header_map[
                display_header
            ]
        )


        return (
            '=IFERROR('
            f'INDEX('
            f"'Vendor Scorecard'!"
            f"${value_column}$5:"
            f"${value_column}${scorecard_last_row},"
            f'MATCH('
            f'$B$3,'
            f"'Vendor Scorecard'!"
            f"${key_column}$5:"
            f"${key_column}${scorecard_last_row},"
            f'0)'
            f'),'
            f'"N/A"'
            f')'
        )


    # ==================================================
    # OVERALL SUMMARY FORMULAS
    # ==================================================

    worksheet["B6"] = (
        lookup_formula(
            "Vendor"
        )
    )

    worksheet["B7"] = (
        lookup_formula(
            "Location"
        )
    )

    worksheet["D6"] = (
        lookup_formula(
            "PO Transactions"
        )
    )

    worksheet["D7"] = (
        lookup_formula(
            "Scored Components"
        )
    )

    worksheet["F6"] = (
        lookup_formula(
            "Prototype Overall Score"
        )
    )

    worksheet["F7"] = (
        lookup_formula(
            "Prototype Overall Grade"
        )
    )

    worksheet["B8"] = (
        lookup_formula(
            "Weight Coverage %"
        )
    )

    worksheet["D8"] = (
        lookup_formula(
            "Overall Status"
        )
    )


    worksheet["F6"].number_format = (
        "0.0"
    )

    worksheet["B8"].number_format = (
        '0.0\\%'
    )


    # ==================================================
    # COMPONENT ROW FORMULAS
    # ==================================================

    for (
        row_number,
        (
            component_name,
            configured_weight,
            metric_header,
            sample_header,
            score_header,
            grade_header,
            status_header
        )
    ) in component_rows.items():

        worksheet.cell(
            row=row_number,
            column=1
        ).value = component_name

        worksheet.cell(
            row=row_number,
            column=2
        ).value = configured_weight

        worksheet.cell(
            row=row_number,
            column=3
        ).value = (
            lookup_formula(
                metric_header
            )
        )

        worksheet.cell(
            row=row_number,
            column=4
        ).value = (
            lookup_formula(
                sample_header
            )
        )

        worksheet.cell(
            row=row_number,
            column=5
        ).value = (
            lookup_formula(
                score_header
            )
        )

        worksheet.cell(
            row=row_number,
            column=6
        ).value = (
            lookup_formula(
                grade_header
            )
        )

        worksheet.cell(
            row=row_number,
            column=7
        ).value = (
            lookup_formula(
                status_header
            )
        )


        for column_number in range(
            1,
            8
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column_number
            )

            cell.border = STANDARD_BORDER

            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )


        worksheet.cell(
            row=row_number,
            column=3
        ).number_format = (
            '0.0\\%'
        )

        worksheet.cell(
            row=row_number,
            column=5
        ).number_format = (
            "0.0"
        )


    # ==================================================
    # SUPPORTING DETAIL FORMULAS
    # ==================================================

    percentage_detail_headers = {
        "OTD %",
        "Supplier-Linked NCR Rate %",
        "NCR Rejected %",
        "Lead-Time Adherence %",
        "NCR Resolution %",
        "Average Price Change %",
        "Price Stability %"
    }


    for (
        start_row,
        section_title,
        metrics
    ) in detail_metrics:

        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=4
        )


        title_cell = worksheet.cell(
            row=start_row,
            column=1
        )

        title_cell.value = (
            section_title
        )

        title_cell.fill = PatternFill(
            "solid",
            fgColor=MEDIUM_BLUE
        )

        title_cell.font = Font(
            color=WHITE,
            bold=True
        )


        for offset, (
            label,
            scorecard_header
        ) in enumerate(
            metrics,
            start=1
        ):

            row_number = (
                start_row
                + offset
            )

            worksheet.cell(
                row=row_number,
                column=1
            ).value = label

            worksheet.cell(
                row=row_number,
                column=1
            ).font = Font(
                bold=True
            )

            worksheet.cell(
                row=row_number,
                column=2
            ).value = (
                lookup_formula(
                    scorecard_header
                )
            )


            worksheet.cell(
                row=row_number,
                column=1
            ).border = STANDARD_BORDER

            worksheet.cell(
                row=row_number,
                column=2
            ).border = STANDARD_BORDER


            if (
                label
                in percentage_detail_headers
            ):

                worksheet.cell(
                    row=row_number,
                    column=2
                ).number_format = (
                    '0.0\\%'
                )


    # ==================================================
    # GRADE CONDITIONAL FORMATTING
    # ==================================================

    grade_cells = [
        "F7",
        "F12",
        "F13",
        "F14",
        "F15",
        "F16"
    ]


    for grade_cell in grade_cells:

        grade_range = (
            worksheet[
                grade_cell
            ]
        )


        grade_range.alignment = Alignment(
            horizontal="center"
        )

        grade_range.font = Font(
            bold=True
        )


    # We use conditional formatting because the values
    # are formulas and change when the dropdown changes.

    grade_range_string = (
        "F7 F12:F16"
    )


    worksheet.conditional_formatting.add(
        grade_range_string,
        CellIsRule(
            operator="equal",
            formula=['"A"'],
            fill=PatternFill(
                "solid",
                fgColor=LIGHT_GREEN
            )
        )
    )

    worksheet.conditional_formatting.add(
        grade_range_string,
        CellIsRule(
            operator="equal",
            formula=['"B"'],
            fill=PatternFill(
                "solid",
                fgColor=LIGHT_BLUE
            )
        )
    )

    worksheet.conditional_formatting.add(
        grade_range_string,
        CellIsRule(
            operator="equal",
            formula=['"C"'],
            fill=PatternFill(
                "solid",
                fgColor=LIGHT_GOLD
            )
        )
    )

    worksheet.conditional_formatting.add(
        grade_range_string,
        CellIsRule(
            operator="equal",
            formula=['"D"'],
            fill=PatternFill(
                "solid",
                fgColor=LIGHT_RED
            )
        )
    )


    # ==================================================
    # NOTES ON DETAIL SHEET
    # ==================================================

    worksheet.merge_cells(
        "A56:G56"
    )

    worksheet["A56"] = (
        "Prototype Interpretation Notes"
    )

    worksheet["A56"].fill = PatternFill(
        "solid",
        fgColor=LIGHT_GOLD
    )

    worksheet["A56"].font = Font(
        bold=True,
        color=GOLD
    )


    detail_notes = [
        (
            "Lead-Time",
            "Currently N/A for reliable scoring because "
            "Item Master benchmark coverage is insufficient."
        ),
        (
            "Responsiveness",
            "NCR resolution percentage is a prototype proxy "
            "and is not actual vendor response time."
        ),
        (
            "Quality",
            "Supplier-linked NCRs are used as a prototype "
            "quality indicator and do not necessarily mean "
            "supplier responsibility has been confirmed."
        ),
        (
            "Commercial",
            "Comparable prices are evaluated only within the "
            "same Vendor + Location + Part + Currency + UOM."
        ),
        (
            "Overall Score",
            "Missing components are excluded and available "
            "weights are normalized. Minimum 3 scored "
            "components are required."
        )
    ]


    for row_offset, (
        topic,
        note
    ) in enumerate(
        detail_notes,
        start=57
    ):

        worksheet.cell(
            row=row_offset,
            column=1
        ).value = topic

        worksheet.cell(
            row=row_offset,
            column=1
        ).font = Font(
            bold=True
        )

        worksheet.merge_cells(
            start_row=row_offset,
            start_column=2,
            end_row=row_offset,
            end_column=7
        )

        worksheet.cell(
            row=row_offset,
            column=2
        ).value = note

        worksheet.cell(
            row=row_offset,
            column=2
        ).alignment = Alignment(
            wrap_text=True
        )


    # ==================================================
    # COLUMN WIDTHS
    # ==================================================

    width_map = {
        "A": 26,
        "B": 22,
        "C": 28,
        "D": 18,
        "E": 18,
        "F": 14,
        "G": 34
    }


    for (
        column_letter,
        width
    ) in width_map.items():

        worksheet.column_dimensions[
            column_letter
        ].width = width


    worksheet.freeze_panes = (
        "A11"
    )

    worksheet.sheet_view.showGridLines = (
        False
    )


# ==================================================
# PROTOTYPE NOTES SHEET
# ==================================================

def _build_prototype_notes_sheet(
    writer,
    rules
):
    """
    Document prototype rules and limitations.
    """

    workbook = writer.book

    worksheet = (
        workbook.create_sheet(
            "Prototype Notes"
        )
    )


    worksheet.merge_cells(
        "A1:D1"
    )

    worksheet["A1"] = (
        "Vendor Scorecard Prototype — Rules & Assumptions"
    )

    worksheet["A1"].fill = PatternFill(
        "solid",
        fgColor=DARK_BLUE
    )

    worksheet["A1"].font = Font(
        color=WHITE,
        bold=True,
        size=15
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center"
    )


    worksheet["A3"] = (
        "IMPORTANT"
    )

    worksheet["B3"] = (
        "This is a working prototype. "
        "Weights, thresholds, minimum samples and grades "
        "have not been approved as production policy."
    )

    worksheet["A3"].fill = PatternFill(
        "solid",
        fgColor=LIGHT_RED
    )

    worksheet["A3"].font = Font(
        bold=True,
        color=RED
    )

    worksheet["B3"].fill = PatternFill(
        "solid",
        fgColor=LIGHT_RED
    )

    worksheet["B3"].alignment = Alignment(
        wrap_text=True
    )


    # ==================================================
    # GRADE THRESHOLDS
    # ==================================================

    worksheet["A5"] = (
        "Prototype Grade Thresholds"
    )

    worksheet["A5"].font = Font(
        bold=True,
        color=DARK_BLUE,
        size=12
    )


    worksheet[
        "A6:B6"
    ][0][0].value = "Grade"

    worksheet[
        "A6:B6"
    ][0][1].value = "Minimum Score"


    for cell in worksheet[
        "6:6"
    ]:

        if cell.column <= 2:

            cell.fill = PatternFill(
                "solid",
                fgColor=DARK_BLUE
            )

            cell.font = Font(
                color=WHITE,
                bold=True
            )


    thresholds = {}

    if rules:

        thresholds = rules.get(
            "grade_thresholds",
            {}
        )


    threshold_order = [
        "A",
        "B",
        "C",
        "D"
    ]


    for index, grade in enumerate(
        threshold_order,
        start=7
    ):

        worksheet.cell(
            row=index,
            column=1
        ).value = grade

        worksheet.cell(
            row=index,
            column=2
        ).value = (
            thresholds.get(
                grade,
                "N/A"
            )
        )

        _apply_grade_style(
            worksheet.cell(
                row=index,
                column=1
            )
        )


    # ==================================================
    # COMPONENT RULES
    # ==================================================

    worksheet["A12"] = (
        "Prototype Component Definitions"
    )

    worksheet["A12"].font = Font(
        bold=True,
        color=DARK_BLUE,
        size=12
    )


    component_headers = [
        "Component",
        "Weight",
        "Prototype Metric / Rule",
        "Important Limitation"
    ]


    for column_index, header in enumerate(
        component_headers,
        start=1
    ):

        cell = worksheet.cell(
            row=13,
            column=column_index
        )

        cell.value = header

        cell.fill = PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )

        cell.font = Font(
            color=WHITE,
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True
        )


    component_definitions = [

        (
            "On-Time Delivery",
            "25%",
            "OTD % based on fully received PO rows with "
            "usable target and last receipt dates.",
            "Target date currently uses Revised Date when "
            "available, otherwise Required Date. "
            "Last Recd Date is aggregate source data."
        ),

        (
            "Quality / NCR",
            "25%",
            "Quality Score = 100 - "
            "(Supplier-Linked NCR Rate % × 5). "
            "NCR Rejected % remains a severity indicator.",
            "Supplier-linked does not necessarily mean "
            "supplier-responsible. PO transaction count is "
            "used as the prototype activity denominator."
        ),

        (
            "Lead-Time Performance",
            "15%",
            "Actual calendar days from PO Date to Last "
            "Receipt Date versus Item Master Lead Time.",
            "Current Item Master benchmark coverage is "
            "insufficient. This component is currently N/A "
            "for reliable vendor scoring."
        ),

        (
            "Responsiveness Proxy",
            "15%",
            "Resolved NCRs / Responsiveness-Eligible NCRs.",
            "This is a prototype proxy only. It is not "
            "actual supplier response time."
        ),

        (
            "Commercial Performance",
            "20%",
            "Price Stability % based on repeat comparable "
            "unit-price purchases.",
            "Comparisons stay within Vendor + Location + "
            "Part + Currency + UOM. It does not account for "
            "quantity breaks or negotiated pricing context."
        )
    ]


    for row_number, component in enumerate(
        component_definitions,
        start=14
    ):

        for column_number, value in enumerate(
            component,
            start=1
        ):

            cell = worksheet.cell(
                row=row_number,
                column=column_number
            )

            cell.value = value

            cell.border = STANDARD_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


    # ==================================================
    # OVERALL SCORE RULE
    # ==================================================

    worksheet["A21"] = (
        "Overall Prototype Score"
    )

    worksheet["A21"].font = Font(
        bold=True,
        color=DARK_BLUE,
        size=12
    )


    worksheet["A22"] = (
        "Minimum scored components"
    )

    minimum_components = (
        rules.get(
            "overall",
            {}
        ).get(
            "minimum_available_components",
            3
        )
        if rules
        else 3
    )

    worksheet["B22"] = (
        minimum_components
    )


    worksheet["A23"] = (
        "Missing component treatment"
    )

    worksheet["B23"] = (
        "Unavailable components are excluded. "
        "Remaining configured weights are normalized "
        "across available scored components."
    )

    worksheet["B23"].alignment = Alignment(
        wrap_text=True
    )


    # ==================================================
    # CURRENT DATA LIMITATIONS
    # ==================================================

    worksheet["A26"] = (
        "Current Prototype Data Limitations"
    )

    worksheet["A26"].font = Font(
        bold=True,
        color=DARK_BLUE,
        size=12
    )


    limitations = [
        "Lead-Time benchmark coverage is currently too low "
        "for meaningful vendor scoring.",
        "Responsiveness uses NCR resolution status rather "
        "than true supplier-response events.",
        "51 supplier-linked NCR records currently cannot be "
        "safely matched to a PO Vendor + Location and remain "
        "in the Unmatched NCRs worksheet.",
        "Vendor Master contains records requiring review, "
        "including incomplete postal codes and exact duplicate "
        "records.",
        "Commercial scoring is a repeat-price prototype and "
        "does not replace formal purchasing or variance analysis."
    ]


    for index, limitation in enumerate(
        limitations,
        start=27
    ):

        worksheet.cell(
            row=index,
            column=1
        ).value = "•"

        worksheet.cell(
            row=index,
            column=2
        ).value = limitation

        worksheet.cell(
            row=index,
            column=2
        ).alignment = Alignment(
            wrap_text=True
        )


    worksheet.column_dimensions[
        "A"
    ].width = 30

    worksheet.column_dimensions[
        "B"
    ].width = 55

    worksheet.column_dimensions[
        "C"
    ].width = 60

    worksheet.column_dimensions[
        "D"
    ].width = 65

    worksheet.sheet_view.showGridLines = (
        False
    )


# ==================================================
# UNMATCHED NCR SHEET
# ==================================================

def _build_unmatched_ncr_sheet(
    writer,
    unmatched_ncrs
):
    """
    Export NCR records that could not be safely matched
    into the operational Vendor + Location scorecard.
    """

    unmatched_export = (
        unmatched_ncrs.copy()
    )


    unmatched_export.to_excel(
        writer,
        sheet_name="Unmatched NCRs",
        index=False
    )


    worksheet = (
        writer.book[
            "Unmatched NCRs"
        ]
    )


    _format_standard_sheet(
        worksheet,
        header_row=1,
        freeze_cell="A2"
    )


    _set_reasonable_widths(
        worksheet,
        maximum_width=34
    )


# ==================================================
# VENDOR REVIEW SHEET
# ==================================================

def _build_vendor_review_sheet(
    writer,
    vendors
):
    """
    Export only Vendor Master records requiring review.
    """

    review_vendors = (
        vendors.loc[
            vendors[
                "review_required"
            ]
        ]
        .copy()
    )


    column_mapping = {

        "company_id":
            "Company ID",

        "vendor_name":
            "Vendor Name",

        "address_line_1":
            "Address",

        "city":
            "City",

        "province":
            "Province / State",

        "postal_code":
            "Postal Code",

        "vendor_quality_status":
            "Vendor Data Status",

        "exact_duplicate_flag":
            "Exact Duplicate",

        "review_reason":
            "Review Reason"
    }


    review_export = (
        _select_existing_columns(
            review_vendors,
            column_mapping
        )
    )


    review_export.to_excel(
        writer,
        sheet_name="Vendor Review",
        index=False
    )


    worksheet = (
        writer.book[
            "Vendor Review"
        ]
    )


    _format_standard_sheet(
        worksheet,
        header_row=1,
        freeze_cell="A2"
    )


    _set_reasonable_widths(
        worksheet,
        maximum_width=34
    )


# ==================================================
# MAIN EXPORT FUNCTION
# ==================================================

def export_vendor_scorecard(
    vendor_summary,
    unmatched_ncrs,
    vendors,
    output_path=(
        "data/output/"
        "Vendor_Scorecard_Prototype.xlsx"
    )
):
    """
    Export the complete Vendor Scorecard prototype.

    Worksheets:
        1. Vendor Scorecard
        2. Vendor Detail
        3. Prototype Notes
        4. Unmatched NCRs
        5. Vendor Review
    """

    output_path = Path(
        output_path
    )


    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )


    rules = (
        _load_scorecard_rules()
    )


    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:


        # ==================================================
        # MAIN SCORECARD
        # ==================================================

        (
            scorecard_sheet,
            scorecard_export
        ) = (
            _build_vendor_scorecard_sheet(
                writer,
                vendor_summary
            )
        )


        # ==================================================
        # INDIVIDUAL VENDOR VIEW
        # ==================================================

        _build_vendor_detail_sheet(
            writer,
            scorecard_export,
            scorecard_sheet
        )


        # ==================================================
        # PROTOTYPE DOCUMENTATION
        # ==================================================

        _build_prototype_notes_sheet(
            writer,
            rules
        )


        # ==================================================
        # EXCEPTIONS
        # ==================================================

        _build_unmatched_ncr_sheet(
            writer,
            unmatched_ncrs
        )


        _build_vendor_review_sheet(
            writer,
            vendors
        )


        # ==================================================
        # WORKBOOK CALCULATION SETTINGS
        # ==================================================
        #
        # Vendor Detail contains formulas that change when
        # a user selects a different Vendor from the dropdown.
        #
        # Ask Excel to recalculate when the workbook opens.
        # ==================================================

        try:

            writer.book.calculation.fullCalcOnLoad = (
                True
            )

            writer.book.calculation.forceFullCalc = (
                True
            )

            writer.book.calculation.calcMode = (
                "auto"
            )

        except AttributeError:

            # Older openpyxl versions may expose calculation
            # properties differently. The workbook still
            # remains valid if these are unavailable.
            pass


    return str(
        output_path
    )